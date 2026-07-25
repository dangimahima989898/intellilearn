import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def get_performance_color(accuracy):
    """
    Returns (bg_hex, text_hex) based on accuracy percentage:
    - Red: < 50%
    - Orange: 50% - 64.99%
    - Yellow: 65% - 74.99%
    - Green: >= 75%
    """
    if accuracy is None:
        return "F3F4F6", "9CA3AF" # Gray fill, gray text for no data
    if accuracy < 50:
        return "FFD2D2", "9C0006" # Soft red, dark red text
    elif accuracy < 65:
        return "FFE0B2", "9C4100" # Soft orange, dark orange text
    elif accuracy < 75:
        return "FFF9C4", "9C8000" # Soft yellow, dark yellow text
    else:
        return "C8E6C9", "1B5E20" # Soft green, dark green text

def generate_heatmap_excel(heatmap_data: list, semester_title: str = "All Semesters") -> bytes:
    """
    Generates a styled Excel sheet for the performance heatmap.
    
    heatmap_data format:
    [
        {
            "subject_code": "CS-101",
            "subject_name": "Computer Networks",
            "units": {
                "Unit 1": {"accuracy": 45.2, "attempts": 20},
                "Unit 2": {"accuracy": 58.0, "attempts": 15},
                ...
            }
        }
    ]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Heatmap"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Arial", size=16, bold=True, color="1F2937")
    subtitle_font = Font(name="Arial", size=11, italic=True, color="4B5563")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    subject_font = Font(name="Arial", size=10, bold=True, color="1F2937")
    regular_font = Font(name="Arial", size=10, color="1F2937")
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Navy Blue
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 1. Title Block
    ws.merge_cells("A1:G1")
    ws["A1"] = "INTELLILEARN ACADEMIC PERFORMANCE HEATMAP"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Course: MCA | Semester Filter: {semester_title} | Generated: Real-Time Admin Export"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = left_align
    ws.row_dimensions[2].height = 20
    
    # Leave row 3 blank
    ws.row_dimensions[3].height = 15
    
    # 2. Table Headers
    headers = ["Subject Code", "Subject Name", "Unit 1", "Unit 2", "Unit 3", "Unit 4", "Unit 5"]
    ws.row_dimensions[4].height = 28
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_idx > 2 else left_align
        cell.border = thin_border
        
    # 3. Populate Data Rows
    current_row = 5
    for item in heatmap_data:
        ws.row_dimensions[current_row].height = 25
        
        # Subject Info
        c_code = ws.cell(row=current_row, column=1, value=item.get("subject_code", ""))
        c_code.font = subject_font
        c_code.alignment = center_align
        c_code.border = thin_border
        
        c_name = ws.cell(row=current_row, column=2, value=item.get("subject_name", ""))
        c_name.font = regular_font
        c_name.alignment = left_align
        c_name.border = thin_border
        
        # Unit Accuracies
        units_dict = item.get("units", {})
        for unit_num in range(1, 6):
            unit_key = f"Unit {unit_num}"
            unit_info = units_dict.get(unit_key, {})
            accuracy = unit_info.get("accuracy")
            attempts = unit_info.get("attempts", 0)
            
            col_idx = 2 + unit_num
            cell = ws.cell(row=current_row, column=col_idx)
            
            if accuracy is not None:
                cell.value = f"{accuracy:.1f}%\n({attempts} att)"
            else:
                cell.value = "N/A\n(0 att)"
                
            bg_color, text_color = get_performance_color(accuracy)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.font = Font(name="Arial", size=9, color=text_color, bold=True)
            cell.alignment = center_align
            cell.border = thin_border
            
        current_row += 1
        
    # 4. Legend Block
    current_row += 2
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    legend_title = ws.cell(row=current_row, column=1, value="Performance Threshold Legend:")
    legend_title.font = Font(name="Arial", size=10, bold=True, color="374151")
    legend_title.alignment = left_align
    
    legend_items = [
        ("< 50% (Critical)", "FFD2D2", "9C0006"),
        ("50% - 64% (Need Attention)", "FFE0B2", "9C4100"),
        ("65% - 74% (Satisfactory)", "FFF9C4", "9C8000"),
        (">= 75% (Excellent)", "C8E6C9", "1B5E20"),
        ("No Data", "F3F4F6", "9CA3AF")
    ]
    
    current_row += 1
    ws.row_dimensions[current_row].height = 20
    for idx, (label, bg, fg) in enumerate(legend_items, 1):
        # Merge first two columns for legend if we need more room, or just set cells
        cell = ws.cell(row=current_row, column=idx if idx < 3 else idx + 1)
        cell.value = label
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.font = Font(name="Arial", size=9, bold=True, color=fg)
        cell.alignment = center_align
        cell.border = thin_border
        
    # Adjust Column Widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    for col in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 16
        
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
